"""
GST Reconciliation Report Generator v3
=======================================
Key Feature: Summary sheet hyperlinked to respective detail sheets.
Clicking any count/amount in Summary navigates directly to that sheet.
"""

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
import warnings
warnings.filterwarnings("ignore")

# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "nav":      "1E3A5F",
    "blue":     "2563EB",
    "matched":  "D1FAE5",
    "probable": "DBEAFE",
    "mismatch": "FEE2E2",
    "only2a":   "FEF3C7",
    "onlybks":  "EDE9FE",
    "white":    "FFFFFF",
    "altrow":   "F8FAFC",
    "total":    "1E3A5F",
    "link":     "2563EB",
    "header_txt": "FFFFFF",
}

def _fill(h):
    return PatternFill("solid", start_color=h, end_color=h)

def _font(bold=False, color="000000", size=10, underline=False):
    return Font(name="Arial", bold=bold, color=color, size=size,
                underline="single" if underline else None)

def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def _border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_border():
    t = Side(style="medium", color="1E3A5F")
    s = Side(style="thin",   color="CBD5E1")
    return Border(left=t, right=t, top=s, bottom=s)

STATUS_FILL = {}


# ── MAIN REPORT WRITER ────────────────────────────────────────────────────────
def write_report(result_df, summary, df_gst, df_books,
                 output_path, period="", source_label="GSTR-2A"):

    wb = Workbook()

    # Resolve dynamic status key
    only_gst_status = f"ONLY IN {source_label.upper()}"
    STATUS_FILL.update({
        "MATCHED":          C["matched"],
        "PROBABLE MATCH":   C["probable"],
        "VENDOR MISMATCH":  C["mismatch"],
        only_gst_status:    C["only2a"],
        "ONLY IN BOOKS":    C["onlybks"],
    })

    # ── Define all detail sheets FIRST so hyperlinks resolve ─────────────────
    ws_all      = wb.create_sheet("Recon Detail")
    ws_matched  = wb.create_sheet("Matched")
    ws_probable = wb.create_sheet("Probable Match")
    ws_mismatch = wb.create_sheet("Vendor Mismatch")
    ws_only_gst = wb.create_sheet(f"Only in {source_label}")
    ws_only_bks = wb.create_sheet("Only in Books")
    ws_vendor   = wb.create_sheet("Vendor Summary")

    # ── Sheet 1: Summary (hyperlinked) ───────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    _title(ws_sum, "GST Reconciliation — Summary Dashboard",
           f"{source_label} vs Books of Accounts", period)

    # Section headers
    ws_sum.row_dimensions[4].height = 8   # spacer

    _section_header(ws_sum, 5, "FINANCIAL OVERVIEW", C["nav"])
    _hrow(ws_sum, 6, ["Particulars", "Amount (Rs)", "Link to Detail"], C["nav"])

    sc         = result_df["Status"].value_counts().to_dict()
    total_gst  = df_gst["Total_Tax"].sum()
    total_bks  = df_books["Total_Tax"].sum()
    net_diff   = total_bks - total_gst
    only_gst_amt = result_df[result_df["Status"]==only_gst_status]["GST_Total_Tax"].sum()
    only_bks_amt = result_df[result_df["Status"]=="ONLY IN BOOKS"]["Books_Total_Tax"].sum()

    fin_rows = [
        (f"Total ITC as per {source_label}",  total_gst,    C["only2a"],   None),
        ("Total ITC as per Books",             total_bks,    C["onlybks"],  None),
        ("Net Difference (Books minus GST)",   net_diff,     C["mismatch"] if net_diff < 0 else C["matched"], None),
    ]
    r = 7
    for label, val, bg, sheet_name in fin_rows:
        _summary_row(ws_sum, r, label, val, bg, sheet_name, is_amount=True)
        r += 1

    ws_sum.row_dimensions[r].height = 8
    r += 1

    # Reconciliation Status Section
    _section_header(ws_sum, r, "RECONCILIATION STATUS", C["nav"])
    r += 1
    _hrow(ws_sum, r, ["Status", "Count", "ITC Amount (Rs)", "Link to Detail"], C["nav"])
    r += 1

    status_rows = [
        ("Matched",          sc.get("MATCHED", 0),
         result_df[result_df["Status"]=="MATCHED"]["GST_Total_Tax"].sum(),
         C["matched"],   "Matched"),
        ("Probable Match",   sc.get("PROBABLE MATCH", 0),
         result_df[result_df["Status"]=="PROBABLE MATCH"]["GST_Total_Tax"].sum(),
         C["probable"],  "Probable Match"),
        ("Vendor Mismatch",  sc.get("VENDOR MISMATCH", 0),
         result_df[result_df["Status"]=="VENDOR MISMATCH"]["GST_Total_Tax"].sum(),
         C["mismatch"],  "Vendor Mismatch"),
        (f"Only in {source_label}", sc.get(only_gst_status, 0),
         only_gst_amt,
         C["only2a"],    f"Only in {source_label}"),
        ("Only in Books",    sc.get("ONLY IN BOOKS", 0),
         only_bks_amt,
         C["onlybks"],   "Only in Books"),
    ]

    for label, count, amount, bg, sheet_name in status_rows:
        _status_summary_row(ws_sum, r, label, count, amount, bg, sheet_name)
        r += 1

    # Totals row
    _totals_row(ws_sum, r, result_df)
    r += 2

    # Vendor Summary link
    _section_header(ws_sum, r, "VENDOR-WISE ANALYSIS", "065F46")
    r += 1
    ws_sum.cell(r, 1, "Click here to view Vendor-wise ITC Summary")
    ws_sum.cell(r, 1).font  = _font(bold=True, color=C["link"], size=11, underline=True)
    ws_sum.cell(r, 1).fill  = _fill("F0FDF4")
    ws_sum.cell(r, 1).border = _border()
    ws_sum.cell(r, 1).alignment = _align("center")
    ws_sum.cell(r, 1).hyperlink = f"#'Vendor Summary'!A1"
    ws_sum.merge_cells(f"A{r}:D{r}")
    ws_sum.row_dimensions[r].height = 24
    r += 2

    # Notes section
    _section_header(ws_sum, r, "IMPORTANT NOTES", "7F1D1D")
    r += 1
    notes = [
        "1. ITC eligibility must be verified against GSTR-2B per Rule 36(4) CGST Rules. GSTR-2A is for reference only.",
        "2. Matching done at GSTIN + Total Tax + Date (±5 days) level as Books contain internal Voucher Numbers.",
        "3. Probable Match records should be verified with supplier before claiming ITC.",
        "4. Only in Books records — follow up with supplier for missing 2A/2B entries.",
        "5. Only in GST records — verify if entry is already booked under a different voucher.",
    ]
    for note in notes:
        ws_sum.cell(r, 1, note).font   = _font(size=9, color="374151")
        ws_sum.cell(r, 1).fill         = _fill("FFF7ED")
        ws_sum.cell(r, 1).border       = _border()
        ws_sum.cell(r, 1).alignment    = _align("left", wrap=True)
        ws_sum.merge_cells(f"A{r}:D{r}")
        ws_sum.row_dimensions[r].height = 18
        r += 1

    # Column widths for summary
    ws_sum.column_dimensions["A"].width = 42
    ws_sum.column_dimensions["B"].width = 18
    ws_sum.column_dimensions["C"].width = 22
    ws_sum.column_dimensions["D"].width = 20
    ws_sum.freeze_panes = "A7"

    # ── Populate detail sheets ────────────────────────────────────────────────
    _detail_sheet(ws_all,      result_df,
                  "All Records — Complete Reconciliation", C["nav"], period, source_label, "Summary")

    _detail_sheet(ws_matched,  result_df[result_df["Status"]=="MATCHED"],
                  "Matched Records", "065F46", period, source_label, "Summary")

    _detail_sheet(ws_probable, result_df[result_df["Status"]=="PROBABLE MATCH"],
                  "Probable Match — Verify with Supplier", "1E40AF", period, source_label, "Summary")

    _detail_sheet(ws_mismatch, result_df[result_df["Status"]=="VENDOR MISMATCH"],
                  "Vendor in Both — Amount Differs", "7F1D1D", period, source_label, "Summary")

    _detail_sheet(ws_only_gst, result_df[result_df["Status"]==only_gst_status],
                  f"In {source_label} — Not in Books", "78350F", period, source_label, "Summary")

    _detail_sheet(ws_only_bks, result_df[result_df["Status"]=="ONLY IN BOOKS"],
                  "In Books — Not in GST Statement", "4C1D95", period, source_label, "Summary")

    # ── Vendor Summary sheet ──────────────────────────────────────────────────
    _vendor_summary_sheet(ws_vendor, result_df, period, source_label)

    wb.save(output_path)
    return output_path


# ── Summary row builders ──────────────────────────────────────────────────────

def _summary_row(ws, r, label, value, bg, sheet_name, is_amount=True):
    c1 = ws.cell(r, 1, label)
    c2 = ws.cell(r, 2, round(value, 2) if isinstance(value, float) else value)
    c1.fill = _fill(bg); c1.font = _font(bold=True, size=10)
    c1.border = _border(); c1.alignment = _align("left")
    c2.fill = _fill(bg); c2.font = _font(bold=True, size=10)
    c2.border = _border(); c2.alignment = _align("right")
    if is_amount: c2.number_format = "#,##0.00"
    ws.merge_cells(f"C{r}:D{r}")
    ws.row_dimensions[r].height = 20


def _status_summary_row(ws, r, label, count, amount, bg, sheet_name):
    """Status row with count, amount AND hyperlink to detail sheet."""
    c1 = ws.cell(r, 1, label)
    c2 = ws.cell(r, 2, count)
    c3 = ws.cell(r, 3, round(amount, 2))
    c4 = ws.cell(r, 4, f"View {label} →")

    for c in [c1, c2, c3, c4]:
        c.fill   = _fill(bg)
        c.border = _border()
        c.font   = _font(size=10)

    c1.font = _font(bold=True, size=10)
    c1.alignment = _align("left")
    c2.alignment = _align("center")
    c2.font = _font(bold=True, size=11)
    c3.alignment = _align("right")
    c3.number_format = "#,##0.00"

    # Hyperlink on View column
    c4.hyperlink  = f"#'{sheet_name}'!A1"
    c4.font       = _font(bold=True, color=C["link"], size=10, underline=True)
    c4.alignment  = _align("center")

    ws.row_dimensions[r].height = 22


def _totals_row(ws, r, result_df):
    total_gst = result_df["GST_Total_Tax"].sum()
    total_bks = result_df["Books_Total_Tax"].sum()
    labels = ["TOTAL", "", f"Rs {total_gst:,.2f}  /  Rs {total_bks:,.2f}", ""]
    for ci, val in enumerate(labels, 1):
        cell = ws.cell(r, ci, val)
        cell.fill   = _fill(C["total"])
        cell.font   = _font(bold=True, color="FFFFFF", size=10)
        cell.border = _border()
        cell.alignment = _align("center" if ci > 1 else "left")
    ws.row_dimensions[r].height = 22


def _section_header(ws, r, title, color):
    ws.merge_cells(f"A{r}:D{r}")
    c = ws.cell(r, 1, f"  {title}")
    c.fill      = _fill(color)
    c.font      = _font(bold=True, color="FFFFFF", size=10)
    c.alignment = _align("left")
    c.border    = _border()
    ws.row_dimensions[r].height = 22


# ── Detail sheet builder ──────────────────────────────────────────────────────

def _detail_sheet(ws, df, subtitle, hdr_color, period,
                  source_label="GSTR-2A", back_sheet="Summary"):

    _title(ws, ws.title, subtitle, period)

    # Back to Summary button row
    ws.merge_cells("A3:D3")
    back = ws.cell(3, 1, "← Back to Summary")
    back.hyperlink  = f"#'{back_sheet}'!A1"
    back.font       = _font(bold=True, color=C["link"], size=10, underline=True)
    back.fill       = _fill("EFF6FF")
    back.alignment  = _align("left")
    back.border     = _border()
    ws.row_dimensions[3].height = 20

    headers = [
        "GSTIN", "Supplier Name", f"Invoice No ({source_label})",
        f"Date ({source_label})", "Date (Books)", "Voucher No (Books)",
        "GST IGST", "GST CGST", "GST SGST", "GST Taxable", "GST Total Tax",
        "Books IGST", "Books CGST", "Books SGST", "Books Total Tax",
        "Difference", "Status",
    ]
    widths = [22, 28, 22, 13, 13, 16, 12, 12, 12, 14, 14, 12, 12, 12, 14, 13, 20]
    _hrow(ws, 4, headers, hdr_color)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if df.empty:
        ws.cell(5, 1, "No records in this category.").font = _font(size=10, color="64748B")
        return

    fields = [
        "GSTIN","Supplier_Name","Invoice_No","Date_GST","Date_Books","Voucher_No",
        "GST_IGST","GST_CGST","GST_SGST","GST_Taxable","GST_Total_Tax",
        "Books_IGST","Books_CGST","Books_SGST","Books_Total_Tax","Difference","Status",
    ]
    num_idx = {6, 7, 8, 9, 10, 11, 12, 13, 14, 15}
    only_gst_status = f"ONLY IN {source_label.upper()}"
    STATUS_FILL[only_gst_status] = C["only2a"]

    for ri, (_, row) in enumerate(df.iterrows(), 5):
        bg = STATUS_FILL.get(str(row.get("Status", "")), C["white"])
        row_bg = bg if ri % 2 == 0 else (
            C["altrow"] if bg == C["white"] else bg
        )
        for ci, field in enumerate(fields, 1):
            val = row.get(field, None)
            if isinstance(val, float) and np.isnan(val): val = None
            if field in ["Date_GST","Date_Books"] and pd.notna(val):
                try:    val = pd.to_datetime(val).strftime("%d-%b-%Y")
                except: pass
            cell = ws.cell(ri, ci, val)
            cell.border    = _border()
            cell.font      = _font(size=9)
            cell.fill      = _fill(bg)
            cell.alignment = _align("right" if (ci-1) in num_idx else "left")
            if (ci-1) in num_idx and val is not None:
                cell.number_format = "#,##0.00"

    # Totals row at bottom
    tr = len(df) + 5
    ws.cell(tr, 1, "TOTAL").font  = _font(bold=True, color="FFFFFF", size=10)
    ws.cell(tr, 1).fill           = _fill(C["total"])
    ws.cell(tr, 1).border         = _border()
    for ci in range(2, len(headers)+1):
        ws.cell(tr, ci).fill   = _fill(C["total"])
        ws.cell(tr, ci).border = _border()

    # Sum numeric columns
    num_col_map = {7:"G", 8:"H", 9:"I", 10:"J", 11:"K",
                   12:"L", 13:"M", 14:"N", 15:"O", 16:"P"}
    for ci, col_letter in num_col_map.items():
        ws.cell(tr, ci).value          = f"=SUM({col_letter}5:{col_letter}{tr-1})"
        ws.cell(tr, ci).number_format  = "#,##0.00"
        ws.cell(tr, ci).font           = _font(bold=True, color="FFFFFF", size=10)
        ws.cell(tr, ci).alignment      = _align("right")

    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"
    ws.freeze_panes    = "A5"


# ── Vendor Summary sheet ──────────────────────────────────────────────────────

def _vendor_summary_sheet(ws, result_df, period, source_label="GSTR-2A"):
    _title(ws, "Vendor Summary", f"Vendor-wise ITC — {source_label} vs Books", period)

    # Back to Summary
    ws.merge_cells("A3:F3")
    back = ws.cell(3, 1, "← Back to Summary")
    back.hyperlink  = "#'Summary'!A1"
    back.font       = _font(bold=True, color=C["link"], size=10, underline=True)
    back.fill       = _fill("EFF6FF")
    back.alignment  = _align("left")
    back.border     = _border()
    ws.row_dimensions[3].height = 20

    headers = [
        "GSTIN", "Supplier Name",
        f"{source_label} ITC (Rs)", "Books ITC (Rs)",
        "Difference (Rs)", "Status",
    ]
    widths = [22, 36, 20, 18, 18, 22]
    _hrow(ws, 4, headers, C["nav"])
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    from gst_recon_engine import build_vendor_summary
    vendor_df = build_vendor_summary(result_df, source_label)

    for ri, (_, vrow) in enumerate(vendor_df.iterrows(), 5):
        diff = vrow["Difference"] or 0
        if abs(diff) <= 10:         bg = C["matched"]
        elif diff < 0:              bg = C["mismatch"]
        else:                       bg = C["onlybks"]

        vals = [
            vrow["GSTIN"], vrow["Supplier_Name"],
            vrow["ITC_GST"], vrow["ITC_Books"], diff, vrow["Status"],
        ]
        for ci, val in enumerate(vals, 1):
            if isinstance(val, float) and np.isnan(val): val = 0
            cell = ws.cell(ri, ci, val)
            cell.border    = _border()
            cell.font      = _font(size=10)
            cell.fill      = _fill(bg if ci == 5 else
                                   (C["altrow"] if ri % 2 == 0 else C["white"]))
            cell.alignment = _align("right" if ci in [3,4,5] else "left")
            if ci in [3, 4, 5]:
                cell.number_format = "#,##0.00"

    # Totals row
    tr = len(vendor_df) + 5
    for ci in range(1, 7):
        cell       = ws.cell(tr, ci)
        cell.fill  = _fill(C["total"])
        cell.font  = _font(bold=True, color="FFFFFF", size=10)
        cell.border = _border()
    ws.cell(tr, 1, "GRAND TOTAL")
    ws.cell(tr, 1).alignment = _align("left")
    ws.cell(tr, 3).value          = f"=SUM(C5:C{tr-1})"
    ws.cell(tr, 4).value          = f"=SUM(D5:D{tr-1})"
    ws.cell(tr, 5).value          = f"=SUM(E5:E{tr-1})"
    for ci in [3, 4, 5]:
        ws.cell(tr, ci).number_format = "#,##0.00"
        ws.cell(tr, ci).alignment     = _align("right")

    ws.auto_filter.ref = f"A4:F4"
    ws.freeze_panes    = "A5"


# ── Common helpers ────────────────────────────────────────────────────────────

def _title(ws, title, subtitle, period=""):
    ws.merge_cells("A1:Q1")
    c = ws["A1"]
    c.value     = title
    c.fill      = _fill(C["nav"])
    c.font      = _font(bold=True, color="FFFFFF", size=14)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:Q2")
    c = ws["A2"]
    c.value     = subtitle + (f"  |  Period: {period}" if period else "")
    c.fill      = _fill(C["blue"])
    c.font      = _font(color="FFFFFF", size=10)
    c.alignment = _align("center")
    ws.row_dimensions[2].height = 18


def _hrow(ws, row_num, values, bg):
    for i, v in enumerate(values, 1):
        c = ws.cell(row_num, i, v)
        c.fill      = _fill(bg)
        c.font      = _font(bold=True, color="FFFFFF", size=10)
        c.alignment = _align("center", wrap=True)
        c.border    = _border()
    ws.row_dimensions[row_num].height = 28